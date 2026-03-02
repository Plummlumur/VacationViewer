"""Shared fixtures for VacationViewer tests."""

import pytest
from datetime import date
from pathlib import Path
from openpyxl import Workbook

from screen.ingest.models import VacationRange


@pytest.fixture
def sample_ranges() -> list[VacationRange]:
    """A small set of vacation ranges for testing."""
    return [
        VacationRange(person_id="P001", start=date(2026, 3, 5), end=date(2026, 3, 10)),
        VacationRange(person_id="P002", start=date(2026, 3, 7), end=date(2026, 3, 14)),
        VacationRange(person_id="P003", start=date(2026, 3, 12), end=date(2026, 3, 18)),
    ]


@pytest.fixture
def default_limits() -> dict[int, int]:
    """Default weekday limits for testing."""
    return {0: 5, 1: 5, 2: 5, 3: 5, 4: 5, 5: 2, 6: 2}


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Create a temporary valid XLSX file for testing."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Person-ID", "Urlaubsstart", "Urlaubsende"])
    ws.append(["P001", date(2026, 3, 5), date(2026, 3, 10)])
    ws.append(["P002", date(2026, 3, 7), date(2026, 3, 14)])
    ws.append(["P003", date(2026, 3, 12), date(2026, 3, 18)])

    file_path: Path = tmp_path / "test_urlaub.xlsx"
    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def invalid_schema_xlsx(tmp_path: Path) -> Path:
    """Create a temporary XLSX with wrong column names."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Von", "Bis"])
    ws.append(["Alice", date(2026, 3, 5), date(2026, 3, 10)])

    file_path: Path = tmp_path / "invalid.xlsx"
    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def bad_dates_xlsx(tmp_path: Path) -> Path:
    """Create a temporary XLSX with some invalid date values."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Person-ID", "Urlaubsstart", "Urlaubsende"])
    ws.append(["P001", date(2026, 3, 5), date(2026, 3, 10)])
    ws.append(["P002", "not-a-date", date(2026, 3, 14)])
    ws.append(["", date(2026, 3, 12), date(2026, 3, 18)])

    file_path: Path = tmp_path / "bad_dates.xlsx"
    wb.save(str(file_path))
    wb.close()
    return file_path
