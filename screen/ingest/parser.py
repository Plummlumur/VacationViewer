"""XLSX parser for vacation data.

Loads an Excel file with columns: Person-ID | Urlaubsstart | Urlaubsende
Validates schema, normalizes dates, and expands ranges into per-day counts.
"""

import logging
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from screen.ingest.models import VacationRange

logger: logging.Logger = logging.getLogger(__name__)

EXPECTED_COLUMNS: list[str] = ["Person-ID", "Urlaubsstart", "Urlaubsende"]


def validate_schema(ws: Worksheet) -> list[str]:
    """Validate that the worksheet has the expected column headers.

    Args:
        ws: The worksheet to validate.

    Returns:
        List of error messages. Empty if valid.
    """
    errors: list[str] = []
    header_row: list[str] = [
        str(cell.value).strip() if cell.value else "" for cell in ws[1]
    ]

    for col in EXPECTED_COLUMNS:
        if col not in header_row:
            errors.append(f"Missing column: '{col}'")

    return errors


def _parse_date(value: object, row_num: int, col_name: str) -> date | None:
    """Parse a cell value into a date object.

    Args:
        value: Cell value (datetime, date, or string).
        row_num: Row number for error reporting.
        col_name: Column name for error reporting.

    Returns:
        Parsed date or None if invalid.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value_stripped: str = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value_stripped, fmt).date()
            except ValueError:
                continue
    logger.warning("Row %d: Invalid date in '%s': %r", row_num, col_name, value)
    return None


def load_xlsx(path: Path) -> list[VacationRange]:
    """Load and parse vacation data from an XLSX file.

    Args:
        path: Path to the Excel file.

    Returns:
        List of validated VacationRange objects.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the schema is invalid.
    """
    if not path.exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws: Worksheet = wb.active  # type: ignore[assignment]

    if ws is None:
        raise ValueError("Workbook has no active worksheet")

    schema_errors: list[str] = validate_schema(ws)
    if schema_errors:
        raise ValueError(f"Schema validation failed: {'; '.join(schema_errors)}")

    # Map column names to indices
    header_row: list[str] = [
        str(cell.value).strip() if cell.value else "" for cell in ws[1]
    ]
    col_idx: dict[str, int] = {name: i for i, name in enumerate(header_row)}

    ranges: list[VacationRange] = []
    row_num: int = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        person_id_raw = row[col_idx["Person-ID"]]
        if person_id_raw is None or str(person_id_raw).strip() == "":
            logger.warning("Row %d: Empty Person-ID, skipping", row_num)
            continue

        person_id: str = str(person_id_raw).strip()

        start_date: date | None = _parse_date(
            row[col_idx["Urlaubsstart"]], row_num, "Urlaubsstart"
        )
        end_date: date | None = _parse_date(
            row[col_idx["Urlaubsende"]], row_num, "Urlaubsende"
        )

        if start_date is None or end_date is None:
            logger.warning("Row %d: Skipping due to invalid dates", row_num)
            continue

        try:
            vacation_range = VacationRange(
                person_id=person_id, start=start_date, end=end_date
            )
            ranges.append(vacation_range)
        except ValueError as e:
            logger.warning("Row %d: %s", row_num, e)

    logger.info(
        "Loaded %d vacation ranges from %d data rows", len(ranges), row_num - 1
    )
    wb.close()
    return ranges


def expand_ranges(ranges: list[VacationRange]) -> dict[date, int]:
    """Expand vacation ranges into per-day vacation counts.

    Each range is expanded day-by-day; overlapping ranges from different
    people are summed up to get the total number of people on vacation
    per day.

    Args:
        ranges: List of VacationRange objects.

    Returns:
        Dictionary mapping each date to the number of people on vacation.
    """
    day_counts: Counter[date] = Counter()

    for vr in ranges:
        current: date = vr.start
        while current <= vr.end:
            day_counts[current] += 1
            current += timedelta(days=1)

    return dict(day_counts)
